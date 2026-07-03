"""Promovidos importer: parse messy multi-sheet XLSX into Registro-ready dicts."""
from __future__ import annotations

import hashlib
import os
import re
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.registro import Registro
from app.services.audit_service import record_audit

_GENERIC_SHEETS = {"C1", "A", "HOJA1", "HOJA 1", "SHEET1"}


def _clean(v) -> str:
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def _edad_from(dia, mes, anio, ref_year: int = 2026) -> Optional[int]:
    try:
        y = int(float(anio))
    except (TypeError, ValueError):
        return None
    if y < 100:  # 2-digit year
        y = 1900 + y if y > 25 else 2000 + y
    if not (1900 <= y <= ref_year):
        return None
    return ref_year - y


def _find_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=8):
        joined = " ".join(_clean(c.value).upper() for c in row)
        if "PRIMER APELLIDO" in joined and "NOMBRE" in joined:
            return row[0].row
    return None


def _file_label(path: str) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    return re.sub(r"_Mayus$", "", base, flags=re.IGNORECASE).strip()


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    estructura = _file_label(path)
    out: list[dict] = []
    for ws in wb.worksheets:
        hdr = _find_header_row(ws)
        if hdr is None:
            continue
        promotor = _clean(ws.title)
        if promotor.upper() in _GENERIC_SHEETS:
            promotor = estructura
        # columns are fixed by the standard template (see spec §5):
        # 2 ap1, 3 ap2, 4 nombre, 5 dia, 6 mes, 7 anio, 8 calle, 9 num,
        # 10 colonia, 11 seccion, 12 telefono
        for row in ws.iter_rows(min_row=hdr + 2, values_only=True):
            row = list(row) + [None] * (12 - len(row))
            ap1, ap2, nombre = _clean(row[1]), _clean(row[2]), _clean(row[3])
            if not (ap1 or ap2 or nombre):
                continue  # empty / spacer row
            nombre_completo = _clean(f"{nombre} {ap1} {ap2}")
            calle, num = _clean(row[7]), _clean(row[8])
            direccion = _clean(f"{calle} {num}") or None
            seccion = _clean(row[10]) or None
            tel = re.sub(r"\D", "", _clean(row[11])) or None
            dia, mes, anio = row[4], row[5], row[6]
            edad = _edad_from(dia, mes, anio)
            nac = "/".join(_clean(x) for x in (dia, mes, anio) if _clean(x))
            observacion = f"nac: {nac}" if nac else None
            out.append({
                "nombre_completo": nombre_completo,
                "direccion": direccion,
                "colonia": _clean(row[9]) or None,
                "seccion": seccion,
                "telefono": tel,
                "edad": edad,
                "observacion": observacion,
                "promotor": promotor,
                "estructura": estructura,
                "_sheet": ws.title,
                "_row": None,  # row index attached by caller for client_uuid
            })
    wb.close()
    return out


def _client_uuid(path: str, sheet: str, idx: int) -> str:
    key = f"{os.path.basename(path)}|{sheet}|{idx}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:32]


def import_rows(db: Session, *, organization_id: str, campaign_id: str, path: str) -> dict:
    """Idempotently import promovidos from ``path`` into ``Registro``.

    Never logs or prints PII — only counts. Writes one ``registro.import``
    AuditLog row per call (batch-level, not per-record).
    """
    rows = parse_workbook(path)
    leidas = len(rows)
    importadas = 0
    duplicadas = 0
    # stable per-sheet counter for deterministic client_uuid
    per_sheet: dict[str, int] = {}
    for r in rows:
        sheet = r["_sheet"]
        idx = per_sheet.get(sheet, 0)
        per_sheet[sheet] = idx + 1
        cuid = _client_uuid(path, sheet, idx)
        exists = db.execute(
            select(Registro.id).where(
                Registro.campaign_id == campaign_id,
                Registro.client_uuid == cuid,
            )
        ).scalar_one_or_none()
        if exists:
            duplicadas += 1
            continue
        db.add(Registro(
            organization_id=organization_id,
            campaign_id=campaign_id,
            activista_id=None,
            nombre_completo=r["nombre_completo"],
            seccion=r["seccion"],
            direccion=r["direccion"],
            colonia=r["colonia"],
            telefono=r["telefono"],
            edad=r["edad"],
            estructura=r["estructura"],
            promotor=r["promotor"],
            observacion=r["observacion"],
            consentimiento=True,
            aviso_version="import-papel-2024",
            client_uuid=cuid,
        ))
        importadas += 1
    file_ref = hashlib.sha1(os.path.basename(path).encode("utf-8")).hexdigest()[:16]
    record_audit(
        db, action="registro.import", actor_id=None,
        organization_id=organization_id, entity_type="registro_batch",
        entity_id=file_ref,
        meta={"leidas": leidas, "importadas": importadas, "duplicadas": duplicadas},
    )
    db.commit()
    return {"leidas": leidas, "importadas": importadas, "duplicadas": duplicadas}
