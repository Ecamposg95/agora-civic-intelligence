"""Parser de Excel de promovidos + import idempotente."""
import openpyxl

from app.services import import_service


def _make_xlsx(path, header_row=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ALAN URIEL RAMIREZ"
    r = header_row
    ws.cell(r, 1, "N.P."); ws.cell(r, 2, "PRIMER APELLIDO"); ws.cell(r, 3, "SEGUNDO APELLIDO")
    ws.cell(r, 4, "NOMBRE"); ws.cell(r, 5, "FECHA DE NACIMIENTO"); ws.cell(r, 8, "DOMICILIO")
    ws.cell(r, 12, "TELÉFONO CON WHATSAPP")
    ws.cell(r+1, 5, "DIA"); ws.cell(r+1, 6, "MES"); ws.cell(r+1, 7, "AÑO")
    ws.cell(r+1, 8, "CALLE"); ws.cell(r+1, 9, "#"); ws.cell(r+1, 10, "BARRIO/COLONIA")
    ws.cell(r+1, 11, "SECCIÓN")
    # data rows
    ws.cell(r+2, 1, 1); ws.cell(r+2, 2, "LEÓN"); ws.cell(r+2, 3, "ALCARAZ"); ws.cell(r+2, 4, "PEDRO")
    ws.cell(r+2, 5, 2); ws.cell(r+2, 6, 3); ws.cell(r+2, 7, 1988)
    ws.cell(r+2, 8, "C. MADERO"); ws.cell(r+2, 9, 506); ws.cell(r+2, 10, "BO. SAN FRANCISCO")
    ws.cell(r+2, 11, 4132); ws.cell(r+2, 12, "7226127261")
    # 2-digit year row
    ws.cell(r+3, 1, 2); ws.cell(r+3, 2, "GONZALEZ"); ws.cell(r+3, 3, "DAVILA"); ws.cell(r+3, 4, "ALBERTO")
    ws.cell(r+3, 5, 3); ws.cell(r+3, 6, 6); ws.cell(r+3, 7, 71)
    ws.cell(r+3, 11, 4130); ws.cell(r+3, 12, "7223478883")
    # empty row (only N.P.)
    ws.cell(r+4, 1, 3)
    wb.save(path)


def test_parse_maps_columns_and_edad(tmp_path):
    p = tmp_path / "ACTIVISMO CULTURA_Mayus.xlsx"
    _make_xlsx(str(p), header_row=1)
    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 2  # empty row skipped
    r0 = rows[0]
    assert r0["nombre_completo"] == "PEDRO LEÓN ALCARAZ"
    assert r0["seccion"] == "4132"
    assert r0["telefono"] == "7226127261"
    assert r0["edad"] == 2026 - 1988
    assert r0["promotor"] == "ALAN URIEL RAMIREZ"
    assert r0["estructura"] == "ACTIVISMO CULTURA"
    assert r0["observacion"].startswith("nac: ")
    assert rows[1]["edad"] == 2026 - 1971  # 2-digit year 71 → 1971


def test_parse_header_on_row_3(tmp_path):
    p = tmp_path / "EMANUEL_Mayus.xlsx"
    _make_xlsx(str(p), header_row=3)
    rows = import_service.parse_workbook(str(p))
    assert len(rows) == 2 and rows[0]["seccion"] == "4132"
